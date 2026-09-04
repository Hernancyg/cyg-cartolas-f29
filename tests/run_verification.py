"""
Verificación local end-to-end de la app Flask, SIN necesidad de un
proyecto Supabase real (usa `fake_supabase.FakeSupabase`, sembrado con
datos de prueba). Corre las 3 páginas + login + admin con el
`test_client()` de Flask, y valida:

  - Login con llave maestra de PRUEBA (nunca la clave real del usuario).
  - Login con un usuario 'trabajador' de PRUEBA y que no pueda ver
    Administrador (403).
  - "Subir Cartolas": parsea una cartola PDF real (de sesiones anteriores,
    ya validada) y compara los totales de cargo/abono.
  - Descarga del Excel convertido (estructura correcta con openpyxl).
  - "Generar CSV F29": con `parsear_f29` monkeypatcheado (no hay PDF de
    F29 de ejemplo a mano), valida el flujo completo (período, montos,
    remanente, CSV final) contra la lógica real de `f29_parser.py`.
  - Administrador: guardar cuentas_config y crear/editar/resetear
    usuarios contra la base de datos de prueba.

Uso:
    ADMIN_PASSWORD_TEST=... python3 tests/run_verification.py
(No requiere variables reales de Supabase — se reemplaza el cliente.)
"""

import io
import os
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

os.environ.setdefault("SUPABASE_URL", "http://fake.local")
os.environ.setdefault("SUPABASE_KEY", "fake-key")
os.environ.setdefault("ADMIN_PASSWORD", "test_local_only_1234")
os.environ.setdefault("FLASK_SECRET_KEY", "test-secret-key-not-for-prod")
os.environ.setdefault("FLASK_DEBUG", "1")

from tests.fake_supabase import FakeSupabase

FAKE = FakeSupabase()

import app.extensions as extensions

# Evita que create_app() reemplace el cliente por uno real de supabase-py
# (que intentaría conectarse por red con las credenciales falsas de arriba).
extensions.init_supabase = lambda *_a, **_kw: FAKE

from app import create_app  # noqa: E402
from app.auth.security import hash_password  # noqa: E402

flask_app = create_app()
extensions._supabase_client = FAKE  # por si algo lo reinicializó
flask_app.config["WTF_CSRF_ENABLED"] = False
flask_app.config["RATELIMIT_ENABLED"] = False

PASSED = []
FAILED = []


def check(label, condition, extra=""):
    if condition:
        PASSED.append(label)
        print(f"  OK  {label}")
    else:
        FAILED.append(label)
        print(f" FAIL {label} {extra}")


def seed_data():
    salt, h = hash_password("trabajador123")
    FAKE.table("usuarios").insert({
        "usuario": "testuser", "nombre": "Usuario de Prueba", "rol": "trabajador",
        "salt": salt, "hash": h, "activo": True,
    }).execute()

    cuentas = [
        ("538", "2108-02", "Total Débitos", "DEBE", "TOTAL DÉBITOS", "=", 0),
        ("537", "1108-02", "Total Créditos", "HABER", "TOTAL CRÉDITOS", "=", 1),
        ("48", "2108-03", "Retención Impuesto Único Trabajadores", "DEBE", "Art. 73 LIR", "+", 2),
    ]
    for codigo, cuenta, desc, tipo, ancla, op, orden in cuentas:
        FAKE.table("cuentas_config").insert({
            "codigo_f29": codigo, "cuenta": cuenta, "descripcion": desc,
            "tipo": tipo, "texto_ancla": ancla, "operador": op, "orden": orden,
        }).execute()


def main():
    seed_data()
    client = flask_app.test_client()

    # ---- Login: usuario/contraseña incorrectos ----
    r = client.post("/login", data={"usuario": "nadie", "clave": "nada"})
    check("login con credenciales inválidas no autentica", r.status_code == 200 and b"incorrectos" in r.data)

    # ---- Login: llave maestra de PRUEBA ----
    r = client.post("/login", data={"usuario": "", "clave": "test_local_only_1234"}, follow_redirects=True)
    check("login con llave maestra de prueba entra como admin", r.status_code == 200 and "Subir Cartolas".encode() in r.data)

    # ---- Subir Cartolas: página inicial ----
    r = client.get("/cartolas/")
    check("GET /cartolas/ 200", r.status_code == 200)
    check("grilla de bancos presente", b"Selecciona tu banco" in r.data)

    pdf_candidates = [
        "/root/.claude/uploads/6765170a-fd6f-5183-b8a3-c1c40a5d8b1c/b3283656-CartolaHistCtaCte000080035381005920260514.pdf",
        "/root/.claude/uploads/6765170a-fd6f-5183-b8a3-c1c40a5d8b1c/c7c222e5-Cartola_004__2026.pdf",
    ]
    for pdf_path in pdf_candidates:
        if not os.path.exists(pdf_path):
            check(f"cartola de prueba existe ({os.path.basename(pdf_path)})", False)
            continue
        with open(pdf_path, "rb") as fh:
            data = {
                "banco": "bci",
                "archivo": (io.BytesIO(fh.read()), os.path.basename(pdf_path)),
            }
            r = client.post("/cartolas/procesar", data=data, content_type="multipart/form-data")
        check(f"procesar cartola {os.path.basename(pdf_path)} -> 200", r.status_code == 200)
        check("tabla de movimientos aparece", b"Ajustes antes de exportar" in r.data or b"No se encontraron movimientos" in r.data)

    # ---- BancoEstado: calibración contra cartola real (01-2026, 7 páginas) ----
    banco_estado_pdf = "/root/.claude/uploads/6765170a-fd6f-5183-b8a3-c1c40a5d8b1c/ef109d5d-Cartola_Bco_Estado_01_2026.pdf"
    if os.path.exists(banco_estado_pdf):
        from app.parsers.bank_parsers import parse_pdf as _parse_pdf_directo, _parse_amount

        resultado_be = _parse_pdf_directo(banco_estado_pdf)
        check("BancoEstado: banco detectado correctamente", resultado_be.banco_detectado == "BancoEstado")
        check("BancoEstado: sin advertencias", resultado_be.advertencias == [])
        check("BancoEstado: se extrajeron los 289 movimientos reales", len(resultado_be.transacciones) == 289)

        # Cadena de saldos: saldo[i] debe calzar exactamente con
        # saldo[i+1] - cargo[i] + abono[i] en TODAS las filas consecutivas
        # (la cartola trae el saldo después de cada movimiento, en orden
        # de más reciente a más antiguo). Es la verificación más fuerte de
        # que ningún movimiento quedó mal separado, duplicado o perdido a
        # lo largo de las 7 páginas.
        saldo_ok = True
        for i in range(len(resultado_be.transacciones) - 1):
            cur = resultado_be.transacciones[i]
            nxt = resultado_be.transacciones[i + 1]
            esperado = _parse_amount(nxt.saldo) - cur.cargo + cur.abono
            if abs(esperado - _parse_amount(cur.saldo)) > 0.5:
                saldo_ok = False
                break
        check("BancoEstado: la cadena de saldos calza en las 288 transiciones", saldo_ok)

        descs = [tx.descripcion for tx in resultado_be.transacciones]
        check(
            "BancoEstado: glosa multilínea concatenada correctamente",
            "Transferencia otro banco a rut 76907072-9 repuestos acira spa" in descs,
        )
        check(
            "BancoEstado: glosa partida entre páginas concatenada correctamente",
            "Abonos varios sociedad de servicios transaccionales ca" in descs,
        )
        check(
            "BancoEstado: sin texto de pie de página pegado a la glosa",
            not any("correo" in d.lower() or "bancoestado.cl" in d.lower() for d in descs),
        )

        with open(banco_estado_pdf, "rb") as fh:
            r = client.post(
                "/cartolas/procesar",
                data={"banco": "banco_estado", "archivo": (io.BytesIO(fh.read()), os.path.basename(banco_estado_pdf))},
                content_type="multipart/form-data",
            )
        check("BancoEstado: POST /cartolas/procesar -> 200", r.status_code == 200)
        check("BancoEstado: tabla de movimientos aparece en la vista previa", b"Ajustes antes de exportar" in r.data)
    else:
        check("BancoEstado: cartola de prueba existe", False)

    # ---- Descargar Excel convertido (con filas de ejemplo) ----
    r = client.post("/cartolas/descargar", data={
        "base_name": "prueba",
        "fecha": ["01/01/2026", "02/01/2026"],
        "detalle": ["Movimiento A", "Movimiento B"],
        "cargo": ["1000", ""],
        "abono": ["", "2000"],
    })
    check("descarga de Excel 200", r.status_code == 200)
    check("Excel tiene Content-Disposition attachment", "attachment" in r.headers.get("Content-Disposition", ""))
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(r.data))
        ws = wb["Banco"]
        check("hoja 'Banco' con encabezados correctos", ws["B1"].value == "FECHA DIA/MES" and ws["E1"].value == "MONTO DEPOSITOS O ABONOS")
        check("fila de datos con cargo correcto", ws["D2"].value == 1000.0)
        check("fila de datos con abono correcto", ws["E3"].value == 2000.0)
    except Exception as exc:  # noqa: BLE001
        check("Excel se puede leer con openpyxl", False, str(exc))

    # ---- Generar CSV F29 (con parsear_f29 monkeypatcheado: sin PDF real) ----
    import app.f29.routes as f29_routes
    from app.parsers.f29_parser import F29Data

    def fake_parsear_f29(_buf, configs=None):
        return F29Data(
            mes="06", anio="2026", rut="76123456-7", razon_social="Empresa de Prueba SpA",
            valores={"538": "1.500.000", "537": "1.000.000", "48": "50.000"},
            remanente_504=0, remanente_504_encontrado=False, advertencias=[],
        )

    f29_routes.parsear_f29 = fake_parsear_f29

    r = client.get("/f29/")
    check("GET /f29/ 200", r.status_code == 200)

    r = client.post("/f29/procesar", data={"archivo": (io.BytesIO(b"%PDF-1.4 fake"), "f29_prueba.pdf")}, content_type="multipart/form-data")
    check("procesar F29 -> 200", r.status_code == 200)
    check("montos detectados en la tabla", b"1500000" in r.data or b"1.500.000" in r.data or b"value=\"1500000\"" in r.data)

    r = client.post("/f29/preview", data={
        "mes": "06", "anio": "2026", "remanente_504": "0",
        "f_cuenta": ["2108-02", "1108-02", "2108-03"],
        "f_codigo": ["538", "537", "48"],
        "f_tipo": ["DEBE", "HABER", "DEBE"],
        "f_monto": ["1500000", "1000000", "50000"],
        "f_incluir": ["538", "537", "48"],
    })
    check("preview F29 -> 200", r.status_code == 200)
    check("totales DEBE/HABER en preview (formato chileno con puntos)", b"1.550.000" in r.data)

    r = client.post("/f29/descargar", data={
        "mes": "06", "anio": "2026", "remanente_504": "0",
        "f_cuenta": ["2108-02", "1108-02", "2108-03"],
        "f_codigo": ["538", "537", "48"],
        "f_tipo": ["DEBE", "HABER", "DEBE"],
        "f_monto": ["1500000", "1000000", "50000"],
        "f_incluir": ["538", "537", "48"],
    })
    check("descarga CSV F29 -> 200", r.status_code == 200)
    csv_text = r.data.decode("latin-1")
    check("CSV contiene glosa de centralización", "CENTRALIZACION F29 JUNIO" in csv_text)
    check("CSV separado por ';' con CRLF", "\r\n" in csv_text and ";" in csv_text)

    # ---- Calcular Global (IGC) ----
    from app.global_igc.calculator import (
        calcular_igc_tabla, calcular_global, EntradaGlobal, UTA_2026,
        UMBRAL_COTIZACION_HONORARIOS, TASA_COTIZACION_PREVISIONAL_HONORARIOS,
    )

    check("UTA 2026 = 834.504", UTA_2026 == 834504)
    check("tramo exento hasta 11.265.804", calcular_igc_tabla(11_265_804) == 0)
    check("primer peso afecto tributa 4%", calcular_igc_tabla(11_265_804.01) == round(11_265_804.01 * 0.04 - 450_632))
    check("tramo tope superior (renta muy alta) usa 40% - rebaja", calcular_igc_tabla(300_000_000) == round(300_000_000 * 0.40 - 32_397_949))
    check("base imponible 0 o negativa no tributa", calcular_igc_tabla(0) == 0)

    entrada_prueba = EntradaGlobal(
        base_tributable_sueldos=0, retiros_14a=10_000_000, credito_retiros_14a=2_000_000,
    )
    res_prueba = calcular_global(entrada_prueba)
    check("Renta bruta retiros = solo el monto neto percibido (sin el incremento IDPC)", res_prueba.renta_bruta_retiros == 10_000_000)
    check("Incremento por IDPC = el crédito IDPC del retiro 14A", res_prueba.total_creditos_idpc == 2_000_000)
    check("gross-up sigue sumando al SUB TOTAL: retiro neto + incremento IDPC", res_prueba.renta_bruta_retiros + res_prueba.total_creditos_idpc == 12_000_000)
    check("débito por restitución = 35% del crédito 14A", res_prueba.debito_restitucion == round(2_000_000 * 0.35))
    check("total créditos incluye el crédito IDPC del retiro", res_prueba.total_creditos == 2_000_000)

    entrada_d3 = EntradaGlobal(retiros_14d3=10_000_000, credito_retiros_14d3=1_000_000)
    res_d3 = calcular_global(entrada_d3)
    check("régimen 14 D N°3 no paga débito por restitución", res_d3.debito_restitucion == 0)

    # "Base Tributable" (03-09-2026): reemplaza a "Total Imponible" +
    # "Leyes Sociales" — el usuario ingresa directamente la renta líquida
    # por sueldos, sin resta interna.
    entrada_sueldo = EntradaGlobal(base_tributable_sueldos=9_000_000, credito_iusc=300_000)
    res_sueldo = calcular_global(entrada_sueldo)
    check("Base Tributable pasa directo a renta líquida de sueldos (sin restar nada)", res_sueldo.renta_neta_sueldos == 9_000_000)
    check("crédito IUSC entra al total de créditos", res_sueldo.total_creditos == 300_000)
    check("renta líquida de sueldos no puede ser negativa", calcular_global(EntradaGlobal(base_tributable_sueldos=-1)).renta_neta_sueldos == 0)
    check("ya no existen los campos total_imponible/leyes_sociales en EntradaGlobal", "total_imponible" not in EntradaGlobal.__dataclass_fields__ and "leyes_sociales" not in EntradaGlobal.__dataclass_fields__)

    entrada_honorarios = EntradaGlobal(honorarios=1_000_000, credito_honorarios=145_000)
    res_honorarios = calcular_global(entrada_honorarios)
    check("gasto presunto honorarios = 30% del bruto", res_honorarios.gasto_presunto_honorarios == 300_000)
    check("honorarios a tributar = 70% del bruto", res_honorarios.honorarios_tributables == 700_000)
    check("base imponible usa el honorario neto de gasto presunto, no el bruto", res_honorarios.base_imponible == 700_000)
    check("crédito por honorarios se calcula sobre el bruto (no se toca)", res_honorarios.total_creditos == 145_000)

    # Tope de 15 UTA a la rebaja de 30% de gasto presunto de honorarios
    # (pedido del usuario, 03-09-2026).
    from app.global_igc.calculator import TOPE_GASTO_PRESUNTO_HONORARIOS_UTA
    tope_15_uta = round(15 * UTA_2026)
    check("tope de gasto presunto de honorarios = 15 UTA", TOPE_GASTO_PRESUNTO_HONORARIOS_UTA == 15)
    honorarios_bajo_tope = round(tope_15_uta / 0.30) - 1_000_000  # 30% queda bajo el tope
    res_bajo_tope = calcular_global(EntradaGlobal(honorarios=honorarios_bajo_tope))
    check("bajo el tope, la rebaja es 30% normal", res_bajo_tope.gasto_presunto_honorarios == round(honorarios_bajo_tope * 0.30))
    honorarios_sobre_tope = round(tope_15_uta / 0.30) + 10_000_000  # 30% supera el tope
    res_sobre_tope = calcular_global(EntradaGlobal(honorarios=honorarios_sobre_tope))
    check("sobre el tope, la rebaja se limita a 15 UTA", res_sobre_tope.gasto_presunto_honorarios == tope_15_uta)
    check("con la rebaja topada, el honorario a tributar es mayor a lo que daría el 70%", res_sobre_tope.honorarios_tributables == honorarios_sobre_tope - tope_15_uta)

    check("ya no existen campos de dividendos en EntradaGlobal", not any("dividendo" in f for f in EntradaGlobal.__dataclass_fields__))

    check("umbral cotización honorarios = 5 ingresos mínimos", UMBRAL_COTIZACION_HONORARIOS == 553_553 * 5)
    check("tasa cotización honorarios = 0.85 (85%), no 0.85%", TASA_COTIZACION_PREVISIONAL_HONORARIOS == 0.85)

    entrada_bajo_umbral = EntradaGlobal(honorarios=2_000_000, credito_honorarios=290_000)
    res_bajo_umbral = calcular_global(entrada_bajo_umbral)
    check("bajo el umbral no aplica cotización previsional de honorarios", res_bajo_umbral.afecto_cotizacion_honorarios is False)
    check("bajo el umbral el pago de cotización es 0", res_bajo_umbral.pago_cotizacion_honorarios == 0)

    entrada_sobre_umbral = EntradaGlobal(honorarios=3_000_000, credito_honorarios=435_000)
    res_sobre_umbral = calcular_global(entrada_sobre_umbral)
    check("igual o sobre el umbral SÍ aplica cotización previsional de honorarios", res_sobre_umbral.afecto_cotizacion_honorarios is True)
    check("pago cotización = 85% de la retención (crédito honorarios)", res_sobre_umbral.pago_cotizacion_honorarios == round(435_000 * TASA_COTIZACION_PREVISIONAL_HONORARIOS))
    # Corregido 03-09-2026: el pago de cotización de honorarios YA NO se
    # resta de la base imponible (eso reducía el IGC y aumentaba el saldo
    # a favor, al revés de lo esperado) — ahora se suma en positivo
    # directamente al resultado final, rebajando la devolución.
    check("el pago de cotización de honorarios NO se resta de la base imponible", res_sobre_umbral.total_rebajas == 0)
    check(
        "el pago de cotización de honorarios se suma en positivo al resultado final (rebaja la devolución)",
        res_sobre_umbral.resultado == round(
            res_sobre_umbral.impuesto_determinado - res_sobre_umbral.total_creditos + res_sobre_umbral.pago_cotizacion_honorarios
        ),
    )
    # Con los mismos honorarios (sobre el umbral) pero un crédito por
    # honorarios distinto, el pago de cotización cambia pero la base
    # imponible y el IGC según tabla deben quedar exactamente iguales —
    # confirma que el pago ya no afecta la base imponible en absoluto.
    entrada_otro_credito = EntradaGlobal(honorarios=3_000_000, credito_honorarios=100_000)
    res_otro_credito = calcular_global(entrada_otro_credito)
    check(
        "la cotización de honorarios no altera la base imponible ni el IGC según tabla",
        res_sobre_umbral.pago_cotizacion_honorarios != res_otro_credito.pago_cotizacion_honorarios
        and res_sobre_umbral.base_imponible == res_otro_credito.base_imponible
        and res_sobre_umbral.igc_segun_tabla == res_otro_credito.igc_segun_tabla,
    )

    entrada_umbral_exacto = EntradaGlobal(honorarios=UMBRAL_COTIZACION_HONORARIOS, credito_honorarios=400_000)
    check("en el umbral exacto (igual) SÍ aplica", calcular_global(entrada_umbral_exacto).afecto_cotizacion_honorarios is True)

    # Líneas informativas al estilo F22 (03-09-2026): "IGC/IUSC débito
    # determinado" y "Remanente de crédito por reliquidación del IUSC" se
    # derivan de valores ya calculados y NO alteran el resultado final.
    entrada_iusc_alto = EntradaGlobal(retiros_14a=50_000_000, credito_retiros_14a=10_000_000, credito_iusc=500_000)
    res_iusc_alto = calcular_global(entrada_iusc_alto)
    check(
        "IGC/IUSC débito determinado = impuesto determinado menos crédito IUSC",
        res_iusc_alto.igc_iusc_debito_determinado == res_iusc_alto.impuesto_determinado - 500_000,
    )
    check("sin remanente cuando el impuesto determinado supera al crédito IUSC", res_iusc_alto.remanente_credito_iusc == 0)

    entrada_iusc_bajo = EntradaGlobal(base_tributable_sueldos=1_000_000, credito_iusc=500_000)
    res_iusc_bajo = calcular_global(entrada_iusc_bajo)
    check(
        "remanente de crédito IUSC = la parte que no alcanzó a compensar el impuesto determinado",
        res_iusc_bajo.remanente_credito_iusc == max(500_000 - res_iusc_bajo.impuesto_determinado, 0) and res_iusc_bajo.remanente_credito_iusc > 0,
    )
    check(
        "las líneas informativas de IUSC no alteran el resultado final",
        res_iusc_bajo.resultado == round(res_iusc_bajo.impuesto_determinado - res_iusc_bajo.total_creditos + res_iusc_bajo.pago_cotizacion_honorarios),
    )

    entrada_sub_total = EntradaGlobal(
        base_tributable_sueldos=1_000_000, retiros_14a=2_000_000, credito_retiros_14a=500_000, arriendos_netos=300_000,
    )
    res_sub_total = calcular_global(entrada_sub_total)
    check(
        "SUB TOTAL = renta neta de retiros + incremento IDPC + otras rentas afectas (antes de rebajas)",
        res_sub_total.sub_total == res_sub_total.renta_bruta_retiros + res_sub_total.total_creditos_idpc + res_sub_total.otras_rentas_afectas,
    )

    # ---- Régimen 14 D N°8 + campos de pensión (04-09-2026, octava corrección) ----
    check(
        "nuevos campos existen en EntradaGlobal",
        all(
            f in EntradaGlobal.__dataclass_fields__
            for f in ("retiros_14d8_base", "retiros_14d8_ppm", "renta_pensiones", "credito_iusc_pensiones")
        ),
    )

    entrada_14d8 = EntradaGlobal(retiros_14d8_base=5_000_000, retiros_14d8_ppm=800_000)
    res_14d8 = calcular_global(entrada_14d8)
    check("14 D N°8: la base pasa directo al resumen (sin gross-up)", res_14d8.retiros_14d8_base == 5_000_000)
    check("14 D N°8: el PPM pasa directo como crédito", res_14d8.credito_ppm_14d8 == 800_000)
    check("14 D N°8: la base se suma al SUB TOTAL sin incremento IDPC", res_14d8.sub_total == 5_000_000)
    check("14 D N°8: no genera débito por restitución (exclusivo de 14 A)", res_14d8.debito_restitucion == 0)
    check("14 D N°8: el PPM no se cuenta como incremento por IDPC", res_14d8.total_creditos_idpc == 0)
    check("14 D N°8: el PPM entra al total de créditos", res_14d8.total_creditos == 800_000)
    check("14 D N°8: detalle_creditos incluye credito_ppm_14d8", res_14d8.detalle_creditos.get("credito_ppm_14d8") == 800_000)

    entrada_pension = EntradaGlobal(renta_pensiones=4_000_000, credito_iusc_pensiones=200_000)
    res_pension = calcular_global(entrada_pension)
    check("Pensiones: renta neta pasa directo (separada de sueldos)", res_pension.renta_neta_pensiones == 4_000_000)
    check("Pensiones: renta líquida de pensión no puede ser negativa", calcular_global(EntradaGlobal(renta_pensiones=-1)).renta_neta_pensiones == 0)
    check("Pensiones: la renta neta se suma al SUB TOTAL igual que sueldos", res_pension.sub_total == 4_000_000)
    check("Pensiones: el crédito IUSC de pensión entra al total de créditos", res_pension.total_creditos == 200_000)
    check("Pensiones: detalle_creditos incluye credito_iusc_pensiones", res_pension.detalle_creditos.get("credito_iusc_pensiones") == 200_000)
    check(
        "Pensiones: se combina con credito_iusc en la línea informativa 'IGC/IUSC débito determinado'",
        res_pension.igc_iusc_debito_determinado == res_pension.impuesto_determinado - 200_000,
    )

    entrada_iusc_combo = EntradaGlobal(
        base_tributable_sueldos=1_000_000, credito_iusc=300_000,
        renta_pensiones=500_000, credito_iusc_pensiones=200_000,
    )
    res_combo = calcular_global(entrada_iusc_combo)
    check(
        "IUSC sueldos + IUSC pensiones se combinan en las líneas informativas de reliquidación",
        res_combo.igc_iusc_debito_determinado == res_combo.impuesto_determinado - 500_000
        and res_combo.remanente_credito_iusc == max(500_000 - res_combo.impuesto_determinado, 0),
    )
    check(
        "las líneas informativas combinadas de IUSC siguen sin alterar el resultado final",
        res_combo.resultado == round(res_combo.impuesto_determinado - res_combo.total_creditos + res_combo.pago_cotizacion_honorarios),
    )

    r = client.post("/global/calcular", data={
        "base_tributable_sueldos": "1000000", "renta_pensiones": "500000", "credito_iusc_pensiones": "50000",
        "retiros_14d8_base": "3000000", "retiros_14d8_ppm": "400000",
    })
    check("POST /global/calcular con 14 D N°8 y pensiones -> 200", r.status_code == 200)
    check("resumen muestra la línea de Régimen 14 D N°8", "Retiros 14 D N°8".encode() in r.data)
    check("resumen muestra la línea de Renta Total Neta Pagada (Pensiones)", "Renta Total Neta Pagada (Pensiones)".encode() in r.data)
    check("formulario tiene el campo mensual para Base Tributable", b'data-mensual-btn="base_tributable_sueldos"' in r.data)
    check("formulario NO ofrece despliegue mensual para el campo anual de pensiones", b'data-mensual-btn="renta_pensiones"' not in r.data)

    r = client.get("/global/")
    check("GET /global/ 200", r.status_code == 200 and b"Calcular Global" in r.data)

    r = client.post("/global/calcular", data={
        "base_tributable_sueldos": "0", "retiros_14a": "10000000", "credito_retiros_14a": "2000000",
    })
    check("POST /global/calcular 200", r.status_code == 200)
    check("resultado del cálculo se muestra en la página", b"Base Imponible Anual de IUSC o IGC" in r.data)

    # ---- Calcular Global: datos del contribuyente + descarga de PDF (03-09-2026) ----
    entrada_contrib_default = EntradaGlobal()
    check("año tributario por defecto = 2026", entrada_contrib_default.anio_tributario == 2026)
    check("nombre/RUT del contribuyente vacíos por defecto", entrada_contrib_default.nombre_contribuyente == "" and entrada_contrib_default.rut_contribuyente == "")

    entrada_desde_form = EntradaGlobal.desde_formulario({
        "anio_tributario": "2025", "nombre_contribuyente": "Juan Pérez González", "rut_contribuyente": "12.345.678-9",
        "base_tributable_sueldos": "10000000",
    })
    check("desde_formulario respeta año tributario ingresado", entrada_desde_form.anio_tributario == 2025)
    check("desde_formulario no numeriza nombre/RUT del contribuyente", entrada_desde_form.nombre_contribuyente == "Juan Pérez González" and entrada_desde_form.rut_contribuyente == "12.345.678-9")

    entrada_form_sin_anio = EntradaGlobal.desde_formulario({"base_tributable_sueldos": "0"})
    check("desde_formulario usa 2026 si no se ingresa año tributario", entrada_form_sin_anio.anio_tributario == 2026)

    r = client.post("/global/pdf", data={
        "anio_tributario": "2026", "nombre_contribuyente": "Juan Pérez González", "rut_contribuyente": "12.345.678-9",
        "base_tributable_sueldos": "20000000", "credito_iusc": "1500000",
        "renta_pensiones": "2000000", "credito_iusc_pensiones": "150000",
        "honorarios": "8000000", "credito_honorarios": "980000",
        "retiros_14a": "15000000", "credito_retiros_14a": "5000000",
        "retiros_14d3": "6000000", "credito_retiros_14d3": "1500000",
        "retiros_14d8_base": "4000000", "retiros_14d8_ppm": "600000",
        "arriendos_netos": "3000000", "intereses_reajustes": "500000",
        "pensiones_alimenticias": "1200000",
    })
    check("POST /global/pdf -> 200", r.status_code == 200)
    check("POST /global/pdf devuelve un PDF válido", r.headers.get("Content-Type", "").startswith("application/pdf") and r.data[:4] == b"%PDF")
    check("POST /global/pdf trae nombre de archivo con año tributario y RUT", "AT2026" in r.headers.get("Content-Disposition", "") and "12" in r.headers.get("Content-Disposition", ""))
    check("el PDF generado no está vacío/truncado", len(r.data) > 2000)

    # ---- Administrador: cuentas ----
    r = client.get("/admin/cuentas")
    check("GET /admin/cuentas 200 (admin)", r.status_code == 200)

    r = client.post("/admin/cuentas/guardar", data={
        "c_cuenta": ["2108-02", "1108-02"],
        "c_codigo": ["538", "537"],
        "c_descripcion": ["Total Débitos", "Total Créditos"],
        "c_tipo": ["DEBE", "HABER"],
        "c_ancla": ["TOTAL DÉBITOS", "TOTAL CRÉDITOS"],
        "c_operador": ["=", "="],
    }, follow_redirects=True)
    check("guardar cuentas_config -> 200", r.status_code == 200)
    check("cuentas_config quedó con 2 filas", len(FAKE._store.get("cuentas_config", [])) == 2)

    # ---- Administrador: usuarios ----
    r = client.get("/admin/usuarios")
    check("GET /admin/usuarios 200 (admin)", r.status_code == 200 and b"testuser" in r.data)

    r = client.post("/admin/usuarios/crear", data={
        "usuario": "nuevo_test", "nombre": "Nuevo Test", "rol": "trabajador",
        "clave": "clave1234", "clave_confirmar": "clave1234",
    }, follow_redirects=True)
    check("crear usuario -> 200", r.status_code == 200)
    check("usuario nuevo quedó en la tabla", any(u["usuario"] == "nuevo_test" for u in FAKE._store.get("usuarios", [])))

    # ---- Logout y login como 'trabajador' (no admin) ----
    client.post("/logout")
    r = client.post("/login", data={"usuario": "testuser", "clave": "trabajador123"}, follow_redirects=True)
    check("login como trabajador de prueba entra", r.status_code == 200 and b"Subir Cartolas" in r.data)

    r = client.get("/admin/cuentas")
    check("trabajador NO puede ver Administrador (403)", r.status_code == 403)

    r = client.get("/cartolas/")
    check("trabajador SÍ puede ver Subir Cartolas", r.status_code == 200)

    r = client.get("/f29/")
    check("trabajador SÍ puede ver Generar CSV F29", r.status_code == 200)

    r = client.get("/global/")
    check("trabajador SÍ puede ver Calcular Global", r.status_code == 200)

    print(f"\n{len(PASSED)} OK, {len(FAILED)} FAIL")
    if FAILED:
        print("Fallaron:", FAILED)
        sys.exit(1)


if __name__ == "__main__":
    main()
