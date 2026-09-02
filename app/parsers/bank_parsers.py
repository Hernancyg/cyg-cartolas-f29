"""
Motor de extracción de movimientos desde cartolas bancarias (PDF o Excel)
de distintos bancos chilenos (Banco de Chile, BancoEstado, Santander, BCI).

Estrategia general para PDF:
  1. Se extraen las palabras de cada página con sus coordenadas (pdfplumber).
  2. Se busca la fila de encabezado de la tabla de movimientos, comparando
     cada palabra contra listas de sinónimos (FECHA, DESCRIPCION, CARGO,
     ABONO, SALDO, etc.).
  3. Se determinan los límites horizontales (x) de cada columna a partir de
     la posición de las palabras del encabezado.
  4. Cada línea siguiente que comienza con una fecha (dd/mm/aaaa o
     dd-mm-aaaa) inicia un movimiento nuevo; las líneas que no comienzan
     con fecha se consideran continuación de la descripción del movimiento
     anterior (glosas que ocupan más de una línea).
  5. Los montos de cada línea se asignan a CARGO/ABONO/SALDO según en qué
     rango de columna caiga su coordenada x0.

Esto es intencionalmente genérico (no hay una función distinta "por banco")
porque las cuatro cartolas objetivo comparten el mismo esquema tabular
(fecha | descripción | cargo | abono | saldo), aunque el texto exacto del
encabezado varía. La detección de banco (detect_bank) es solo informativa /
para mostrar en la UI, no cambia la lógica de extracción.

Validado con una cartola real de BCI. Para BancoEstado, Santander y Banco de
Chile la lógica es la misma pero no ha sido probada contra un PDF real de
esos bancos — si el resultado no calza, conviene revisar los sinónimos de
encabezado más abajo o enviar un PDF de ejemplo para calibrar.
"""

import re
from dataclasses import dataclass, field

import pdfplumber


# ---------------------------------------------------------------------------
# Sinónimos de encabezado por rol de columna
# ---------------------------------------------------------------------------

HEADER_SYNONYMS = {
    "fecha": ["FECHA"],
    "descripcion": [
        "DESCRIPCION", "DESCRIPCIÓN", "DETALLE", "GLOSA", "MOVIMIENTO",
        "TRANSACCION", "TRANSACCIÓN",
    ],
    "cargo": [
        "CARGOS", "CARGO", "CHEQUES", "DEBITO", "DÉBITO", "DEBITOS",
        "DÉBITOS",
    ],
    "abono": [
        "ABONOS", "ABONO", "DEPOSITOS", "DEPÓSITOS", "DEPOSITO",
        "DEPÓSITO", "CREDITO", "CRÉDITO", "CREDITOS", "CRÉDITOS",
    ],
    "saldo": ["SALDO"],
    "documento": [
        "DOCUMENTO", "DOCUMENTOS", "DOC", "N°", "Nº", "Nª", "N°DOC",
        "OPERACION", "OPERACIÓN",
    ],
    "sucursal": ["SUCURSAL"],
}

# Palabras que marcan el fin de la tabla de movimientos (resumen / letra legal)
SECTION_END_MARKERS = ["RESUMEN", "CONSIDERAMOS"]

# Frases de dos o más palabras que marcan el inicio de una tabla distinta
# (p.ej. "Saldos diarios" en BancoEstado) que NO debe procesarse como si
# fueran más movimientos: una vez encontrada, se corta ahí mismo y se
# ignora el resto del documento (la tabla de saldos diarios suele seguir
# ocupando varias páginas más).
HARD_STOP_PHRASES = [["SALDOS", "DIARIOS"]]

# Frases de dos o más palabras del pie de página legal/nota que se repite
# al final de cada página (p.ej. "Nota: Información..." en BancoEstado,
# "Retención a 1 día..." en Banco de Chile). A diferencia de
# HARD_STOP_PHRASES, esto solo recorta la página actual (la tabla de
# movimientos puede seguir en la página siguiente); evita que el bloque de
# retenciones/resumen o el número de página quede pegado al último
# movimiento de la página.
FOOTER_STOP_PHRASES = [
    ["NOTA", "INFORMACIÓN"], ["NOTA", "INFORMACION"],
    ["RETENCION", "A", "1", "DIA"], ["RETENCIÓN", "A", "1", "DIA"],
    # Pie de página legal repetido en cada hoja de BancoEstado ("Para
    # solicitud de requerimientos... enviar correo a: ..." / "Para
    # canalizar sus solicitudes relacionadas con Internet, enviar correo
    # a: ..."). Sin este corte, ese texto queda pegado a la glosa del
    # último movimiento de la página.
    ["PARA", "SOLICITUD"], ["PARA", "CANALIZAR"],
]

# Descripciones que marcan una fila de saldo (no un movimiento real) y que
# por lo tanto no deben incluirse en la salida — p.ej. "SALDO INICIAL" /
# "SALDO FINAL" en Banco de Chile.
SALDO_ROW_MARKERS = ("SALDO INICIAL", "SALDO FINAL")


def _is_saldo_row(descripcion: str) -> bool:
    d = " ".join((descripcion or "").upper().split())
    return any(d.startswith(marker) for marker in SALDO_ROW_MARKERS)


# Algunas cartolas (p.ej. Banco de Chile) no incluyen el año en la fecha de
# cada movimiento ("02/01"), solo día/mes — el año se obtiene del período
# de la cartola ("DESDE : 30/12/2025 HASTA : 30/01/2026", presente en el
# encabezado de cada página).
PERIOD_RE = re.compile(
    r"DESDE\s*:?\s*\d{1,2}/(\d{1,2})/(\d{4})\s+HASTA\s*:?\s*\d{1,2}/(\d{1,2})/(\d{4})"
)


def _extract_month_year_map(full_text: str) -> dict:
    """Busca 'DESDE dd/mm/aaaa HASTA dd/mm/aaaa' en el texto completo del
    PDF y devuelve {mes: año} para cada mes del período (soporta que el
    período cruce de un año al siguiente, p.ej. diciembre -> enero)."""
    m = PERIOD_RE.search(full_text.upper())
    if not m:
        return {}
    mes_desde, anio_desde, mes_hasta, anio_hasta = (int(g) for g in m.groups())
    mapping = {}
    mes, anio = mes_desde, anio_desde
    for _ in range(24):  # tope de seguridad; ninguna cartola real dura 2 años
        mapping[mes] = anio
        if (mes, anio) == (mes_hasta, anio_hasta):
            break
        mes += 1
        if mes > 12:
            mes = 1
            anio += 1
    return mapping


def _normalize_fecha(fecha_str: str, month_year_map: dict):
    """Si `fecha_str` ya tiene año (dd/mm/aaaa), la deja igual. Si solo
    tiene día/mes (Banco de Chile), le agrega el año según
    `month_year_map`. Si no hay año disponible para ese mes, devuelve el
    string original sin modificar (no se adivina un año)."""
    partes = re.split(r"[/-]", fecha_str)
    if len(partes) != 2:
        return fecha_str, True
    dia, mes = partes
    try:
        mes_int = int(mes)
    except ValueError:
        return fecha_str, True
    anio = month_year_map.get(mes_int)
    if anio is None:
        return fecha_str, False
    return f"{dia}/{mes}/{anio}", True


def _find_phrase_top(rows, phrases):
    """Busca, en las filas de una página, alguna de las frases dadas
    (listas de 2+ palabras consecutivas) y devuelve el 'top' de la fila
    donde aparece, o None si no se encontró."""
    for row in rows:
        tokens = [w["text"].upper().rstrip(":") for w in row]
        for phrase in phrases:
            n = len(phrase)
            for i in range(len(tokens) - n + 1):
                if tokens[i:i + n] == phrase:
                    return row[0]["top"]
    return None

BANK_MARKERS = {
    "BCI": ["BCI", "BANCO CREDITO E INVERSIONES", "BANCO CRÉDITO E INVERSIONES", "BANCO CREDITO INVERSIONES"],
    "Banco de Chile": ["BANCO DE CHILE"],
    "BancoEstado": ["BANCOESTADO", "BANCO ESTADO", "BANCO DEL ESTADO"],
    "Santander": [
        "SANTANDER",
        "CARTOLAS HISTÓRICAS DE CTA.CTE", "CARTOLAS HISTORICAS DE CTA.CTE",
    ],
}

DATE_RE = re.compile(r"^\d{1,2}[/-]\d{1,2}(?:[/-]\d{2,4})?$")
# Algunas cartolas (p.ej. BancoEstado) anteponen el signo peso ("$5.000.000")
# a cada monto de la tabla; el "$" es opcional para no afectar bancos que no
# lo usan.
AMOUNT_RE = re.compile(r"^-?\$?\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{1,2})?$")
ROW_Y_TOLERANCE = 3.0  # puntos de tolerancia para agrupar palabras en una misma fila


@dataclass
class Transaction:
    fecha: str = ""
    descripcion: str = ""
    cargo: float = 0.0
    abono: float = 0.0
    saldo: str = ""  # informativo, no se usa en la plantilla de salida


@dataclass
class ParseResult:
    banco_detectado: str = "Desconocido"
    transacciones: list = field(default_factory=list)
    advertencias: list = field(default_factory=list)


def detect_bank(text_upper: str) -> str:
    for bank, markers in BANK_MARKERS.items():
        for marker in markers:
            if marker in text_upper:
                return bank
    return "Desconocido"


def _parse_amount(token: str) -> float:
    """Convierte '1.234.567' o '1,234,567.89' o '1234,50' a float. Admite un
    signo peso opcional al inicio ('$1.234.567', usado por BancoEstado)."""
    t = token.strip().replace("$", "")
    negative = t.startswith("-")
    t = t.lstrip("-")
    # Formato chileno: punto = miles, coma = decimal (si aparece coma al final)
    if "," in t and t.rfind(",") > t.rfind("."):
        t = t.replace(".", "").replace(",", ".")
    else:
        t = t.replace(",", "").replace(".", "")
        # si quedó un solo bloque de dígitos, no hay decimales que perder
    try:
        value = float(t)
    except ValueError:
        return 0.0
    return -value if negative else value


def _group_words_into_rows(words):
    """Agrupa palabras (con x0/x1/top) en filas según su coordenada 'top'."""
    rows = []
    current_row = []
    current_top = None
    for w in sorted(words, key=lambda w: (round(w["top"], 1), w["x0"])):
        if current_top is None or abs(w["top"] - current_top) <= ROW_Y_TOLERANCE:
            current_row.append(w)
            current_top = w["top"] if current_top is None else current_top
        else:
            rows.append(current_row)
            current_row = [w]
            current_top = w["top"]
    if current_row:
        rows.append(current_row)
    return rows


def _match_roles_in_row(row):
    """Devuelve {rol: [palabras]} para las palabras de una fila que calzan
    con algún sinónimo de columna.

    Algunas cartolas (p.ej. BancoEstado) imprimen el encabezado de dos
    columnas pegado en una sola palabra, unido por "/" y sin espacios
    ("Cheques/Cargos", "Depósitos/Abonos"). Para reconocerlas igual, además
    de comparar la palabra completa se compara cada parte separada por "/".
    """
    role_words = {}
    for w in row:
        raw = w["text"].upper().strip(":")
        candidates = {raw.strip("°").strip()}
        if "/" in raw:
            candidates.update(part.strip("°").strip() for part in raw.split("/") if part.strip())
        matched_roles = set()
        for token in candidates:
            for role, synonyms in HEADER_SYNONYMS.items():
                if token in synonyms:
                    matched_roles.add(role)
        for role in matched_roles:
            role_words.setdefault(role, []).append(w)
    return role_words


def _find_header_columns(rows):
    """
    Busca, entre las primeras filas de la página, la fila que contiene las
    palabras clave de columna, y devuelve un dict rol -> (x0, x1) con el
    rango horizontal de esa columna.

    Algunas cartolas tienen, ANTES de la tabla de movimientos, un bloque de
    resumen ("Saldo inicial: ... Depósitos: ... Otros abonos: ...", "Cheques:
    ... Otros cargos: ... Saldo final: ...") cuyas palabras coinciden con los
    mismos sinónimos de columna (CARGO, ABONO, SALDO) pero en posiciones muy
    distintas. Para no confundir esas líneas con el encabezado real, se
    exige que la fila candidata a encabezado contenga "fecha" junto con al
    menos "cargo" o "abono" en la MISMA fila (la línea de resumen no tiene
    la palabra "Fecha"). Si ninguna fila cumple ese criterio más estricto,
    se cae de vuelta al comportamiento anterior (agregar todas las filas
    que matchean algo), por seguridad.
    """
    # Buscamos hasta la primera fila que empiece con una fecha Y además
    # tenga algún monto en esa misma fila (ahí ya comienzan los
    # movimientos), con un tope de seguridad de 40 filas. Se exige el monto
    # además de la fecha porque algunas cartolas (p.ej. BancoEstado) tienen,
    # antes de la tabla, un timbre "Fecha - Hora: 20/07/2026 - 00:35" cuya
    # primera palabra también matchea el formato de fecha, pero que no es
    # todavía una fila de movimiento — sin este chequeo adicional, cortaría
    # la búsqueda del encabezado antes de llegar a la fila real.
    def _row_tiene_monto(row):
        return any(
            AMOUNT_RE.match(w["text"]) and any(ch.isdigit() for ch in w["text"])
            for w in row[1:]
        )

    search_limit = len(rows)
    for idx, row in enumerate(rows[:40]):
        if row and DATE_RE.match(row[0]["text"]) and _row_tiene_monto(row):
            search_limit = idx
            break
    else:
        search_limit = min(40, len(rows))

    per_row_roles = [_match_roles_in_row(row) for row in rows[:search_limit]]

    strong_candidates = [
        idx for idx, roles in enumerate(per_row_roles)
        if "fecha" in roles and ("cargo" in roles or "abono" in roles)
    ]

    if strong_candidates:
        header_idx = max(strong_candidates)
        role_words = {role: list(ws) for role, ws in per_row_roles[header_idx].items()}
        header_top = rows[header_idx][0]["top"] if rows[header_idx] else 0

        # Enriquecer con filas vecinas MUY cercanas verticalmente (encabezado
        # partido en dos líneas), pero solo si esa fila vecina no es a su vez
        # una fila "fuerte" independiente (para no mezclar con otra sección).
        for neighbor_idx in (header_idx - 1, header_idx + 1):
            if neighbor_idx < 0 or neighbor_idx >= len(per_row_roles):
                continue
            if neighbor_idx in strong_candidates:
                continue
            neighbor_row = rows[neighbor_idx]
            if not neighbor_row:
                continue
            neighbor_top = neighbor_row[0]["top"]
            if abs(neighbor_top - header_top) > 15:
                continue
            for role, ws in per_row_roles[neighbor_idx].items():
                role_words.setdefault(role, [])
                role_words[role].extend(ws)
    else:
        # Comportamiento anterior (menos estricto) como red de seguridad.
        role_words = {}
        header_row_indices = []
        for idx, roles in enumerate(per_row_roles):
            if roles:
                header_row_indices.append(idx)
            for role, ws in roles.items():
                role_words.setdefault(role, []).extend(ws)
        header_idx = max(header_row_indices) if header_row_indices else 0

    if "fecha" not in role_words or "cargo" not in role_words or "abono" not in role_words:
        return None, None

    columns = {}
    for role, ws in role_words.items():
        x0 = min(w["x0"] for w in ws)
        x1 = max(w["x1"] for w in ws)
        columns[role] = (x0, x1)

    return columns, header_idx


AMOUNT_COLUMN_TOLERANCE = 20.0  # puntos de margen fuera del rango de la columna


def _classify_amount_column(x0, columns):
    """Devuelve 'cargo', 'abono' o 'saldo' según en qué columna cae x0.
    Si x0 está lejos de las tres columnas (más allá de
    AMOUNT_COLUMN_TOLERANCE puntos del rango de cada una), devuelve None:
    un token con forma de número que cae muy lejos de esas columnas es más
    probable que sea parte de la glosa (p.ej. un código de comercio como
    'Compra CV 371') que un monto real."""
    ranges = {role: c for role, c in columns.items() if role in ("cargo", "abono", "saldo")}
    if not ranges:
        return None

    def distance(rng):
        lo, hi = rng
        if lo <= x0 <= hi:
            return 0.0
        return min(abs(x0 - lo), abs(x0 - hi))

    best_role, best_dist = min(
        ((role, distance(rng)) for role, rng in ranges.items()),
        key=lambda kv: kv[1],
    )
    if best_dist > AMOUNT_COLUMN_TOLERANCE:
        return None
    return best_role


def _find_cutoff_top(words):
    """Devuelve el 'top' más pequeño entre las palabras que marcan el fin de
    la tabla de movimientos (p.ej. 'Resumen del Periodo'), o None si no hay."""
    tops = [
        w["top"] for w in words
        if w["text"].upper().rstrip(":") in SECTION_END_MARKERS
    ]
    return min(tops) if tops else None


def _find_hard_stop_top(rows):
    """Busca, en las filas de una página, alguna de las frases de
    HARD_STOP_PHRASES (p.ej. 'Saldos diarios') que marca el inicio de una
    tabla completamente distinta a la de movimientos. Devuelve el 'top' de
    la fila donde aparece, o None si no se encontró."""
    return _find_phrase_top(rows, HARD_STOP_PHRASES)


def _desc_bounds(columns):
    """Calcula el rango horizontal [x0, x1) de la columna de descripción a
    partir de sus columnas vecinas.

    El orden de las columnas (fecha/cargo/abono/descripción/saldo/
    documento/sucursal) varía según el banco: en algunos la descripción va
    justo después de la fecha, en otros va después de cargo/abono, o
    incluso después del saldo. Para no asumir un orden fijo, se ubican las
    columnas vecinas de "descripción" por posición horizontal (la que
    quede inmediatamente a su izquierda y la que quede inmediatamente a su
    derecha, según el centro de cada columna), y se usa el borde entre
    ellas como límite real de la descripción. El contenido de la columna
    descripción suele partir alineado más a la izquierda que su propia
    etiqueta de encabezado (que puede estar centrada en una columna
    ancha), por eso se usa el borde de la columna vecina y no el x0 del
    encabezado "DESCRIPCION" mismo.
    """
    fecha_x0, fecha_x1 = columns["fecha"]
    other_columns = [(role, c) for role, c in columns.items() if role != "descripcion"]
    if "descripcion" in columns:
        desc_center = sum(columns["descripcion"]) / 2
    else:
        desc_center = fecha_x1  # sin columna de descripción detectada: usar fallback

    left_neighbors = [c[1] for _, c in other_columns if (c[0] + c[1]) / 2 < desc_center]
    right_neighbors = [c[0] for _, c in other_columns if (c[0] + c[1]) / 2 > desc_center]
    desc_x0 = max(left_neighbors) if left_neighbors else fecha_x1
    desc_x1_bound = min(right_neighbors) if right_neighbors else float("inf")
    return desc_x0, desc_x1_bound


def _date_words_in_column(words, columns):
    """Devuelve las palabras de `words` que caen en la columna de fecha,
    con tolerancia amplia y simétrica: en algunas cartolas (p.ej.
    BancoEstado) el texto de la columna "Fecha" queda más a la izquierda
    que la propia etiqueta del encabezado "Fecha" (que puede estar
    centrada en una columna más ancha), por lo que un margen chico solo
    hacia la derecha (como el usado para otras columnas) descarta las
    fechas reales."""
    fecha_x0, fecha_x1 = columns["fecha"]
    return [
        w for w in words
        if DATE_RE.match(w["text"]) and fecha_x0 - 20 <= w["x0"] <= fecha_x1 + 20
    ]


def _process_page_transactions(words, columns):
    """
    Dado el listado de palabras de UNA página (ya recortado a la zona de
    datos de la tabla) y el dict de columnas, devuelve una lista de
    Transaction. Usa la heurística de "fecha más cercana" para asignar cada
    palabra a su movimiento, porque en varios bancos el bloque de la
    descripción queda verticalmente centrado respecto a la fila que
    contiene fecha/monto/saldo (una línea de la glosa puede quedar *arriba*
    de esa fila y otra *abajo*).
    """
    documento_range = columns.get("documento")
    desc_x0, desc_x1_bound = _desc_bounds(columns)

    date_words = _date_words_in_column(words, columns)
    if not date_words:
        return []

    date_tops = [w["top"] for w in date_words]
    txs = [Transaction(fecha=w["text"]) for w in date_words]
    desc_words_per_tx = [[] for _ in txs]

    date_word_ids = {id(w) for w in date_words}

    for w in words:
        if id(w) in date_word_ids:
            continue
        token = w["text"]
        # movimiento más cercano verticalmente
        idx = min(range(len(date_tops)), key=lambda i: abs(date_tops[i] - w["top"]))
        tx = txs[idx]

        in_documento_col = (
            documento_range is not None
            and documento_range[0] - 5 <= w["x0"] < documento_range[1] + 5
        )
        if in_documento_col:
            # N° de documento/cheque: no es un monto ni parte de la glosa
            continue
        role = None
        if AMOUNT_RE.match(token) and any(ch.isdigit() for ch in token):
            role = _classify_amount_column(w["x0"], columns)

        if role is not None:
            value = _parse_amount(token)
            if role == "cargo":
                tx.cargo += value
            elif role == "abono":
                tx.abono += value
            elif role == "saldo":
                tx.saldo = token
        elif (
            any(ch.isalnum() for ch in token)
            and desc_x0 <= w["x0"] < desc_x1_bound
        ):
            # Un token con forma de número que cayó lejos de las columnas
            # de montos (role is None) se trata como parte de la glosa,
            # p.ej. un código de comercio como "Compra CV 371".
            desc_words_per_tx[idx].append(token)

    for tx, dws in zip(txs, desc_words_per_tx):
        tx.descripcion = " ".join(dws).strip()

    return txs


def parse_pdf(file_path_or_buffer) -> ParseResult:
    result = ParseResult()
    transactions = []

    with pdfplumber.open(file_path_or_buffer) as pdf:
        # detectar banco usando el texto completo
        full_text = "\n".join((p.extract_text() or "") for p in pdf.pages)
        result.banco_detectado = detect_bank(full_text.upper())
        month_year_map = _extract_month_year_map(full_text)

        columns = None

        for page in pdf.pages:
            words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
            if not words:
                continue
            rows = _group_words_into_rows(words)

            page_columns, header_idx = _find_header_columns(rows)
            if page_columns:
                columns = page_columns
                header_bottom = max(w["bottom"] for w in rows[header_idx])
            else:
                header_bottom = 0

            if columns is None:
                continue

            cutoff_top = _find_cutoff_top(words)
            footer_top = _find_phrase_top(rows, FOOTER_STOP_PHRASES)
            hard_stop_top = _find_hard_stop_top(rows)
            upper_bounds = [
                t for t in (cutoff_top, footer_top, hard_stop_top) if t is not None
            ]
            upper_bound = min(upper_bounds) if upper_bounds else None
            data_words = [
                w for w in words
                if w["top"] > header_bottom
                and (upper_bound is None or w["top"] < upper_bound)
            ]

            # Continuación de glosa entre páginas: si la última línea de la
            # descripción de un movimiento queda justo en el borde inferior
            # de una página, esa línea puede aparecer recién al comienzo de
            # la página siguiente, ANTES de la fecha del primer movimiento
            # de esa página (p.ej. "transaccionales ca" en una cartola real
            # de BancoEstado). Sin este ajuste, la heurística de "fecha más
            # cercana" la asignaría erróneamente al primer movimiento de la
            # página nueva en lugar de al último de la página anterior.
            date_words_page = _date_words_in_column(data_words, columns)
            if date_words_page and transactions:
                first_date_top = min(w["top"] for w in date_words_page)
                desc_x0, desc_x1_bound = _desc_bounds(columns)
                orphan_words = [
                    w for w in data_words
                    if w["top"] < first_date_top
                    and any(ch.isalnum() for ch in w["text"])
                    and desc_x0 <= w["x0"] < desc_x1_bound
                    and not (AMOUNT_RE.match(w["text"]) and any(ch.isdigit() for ch in w["text"]))
                ]
                if orphan_words:
                    extra = " ".join(w["text"] for w in orphan_words).strip()
                    if extra:
                        transactions[-1].descripcion = (
                            transactions[-1].descripcion + " " + extra
                        ).strip()
                    orphan_ids = {id(w) for w in orphan_words}
                    data_words = [w for w in data_words if id(w) not in orphan_ids]

            transactions.extend(_process_page_transactions(data_words, columns))

            if hard_stop_top is not None:
                # A partir de aquí la(s) página(s) restantes son otra tabla
                # (p.ej. "Saldos diarios"), no más movimientos: se detiene
                # el procesamiento por completo.
                break

    # Filtrar filas de saldo (no son movimientos reales) y completar el año
    # de las fechas que vinieron sin él (p.ej. Banco de Chile).
    transactions = [tx for tx in transactions if not _is_saldo_row(tx.descripcion)]
    for tx in transactions:
        tx.fecha, ok = _normalize_fecha(tx.fecha, month_year_map)
        if not ok:
            result.advertencias.append(
                f"No se pudo determinar el año de la fecha '{tx.fecha}' "
                "(no se encontró el período DESDE/HASTA en el PDF). Revisa "
                "esa fila antes de descargar."
            )

    if not transactions:
        result.advertencias.append(
            "No se detectaron movimientos. Puede que el formato del PDF no "
            "coincida con los encabezados esperados (Fecha/Descripción/Cargo/"
            "Abono). Revisa el archivo o avisa para calibrar el parser con "
            "este banco."
        )

    result.transacciones = transactions
    return result


# ---------------------------------------------------------------------------
# Excel de origen (cartola ya exportada a Excel por el banco)
# ---------------------------------------------------------------------------

def parse_excel(file_path_or_buffer) -> ParseResult:
    import openpyxl

    result = ParseResult()
    wb = openpyxl.load_workbook(file_path_or_buffer, data_only=True)

    all_text_upper = ""
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            all_text_upper += " ".join(str(c) for c in row if c is not None).upper() + "\n"
    result.banco_detectado = detect_bank(all_text_upper)

    transactions = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        header_row_idx = None
        col_map = {}
        for idx, row in enumerate(rows):
            local_map = {}
            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                token = str(cell).upper().strip(":").strip()
                for role, synonyms in HEADER_SYNONYMS.items():
                    if token in synonyms:
                        local_map[role] = col_idx
            if "fecha" in local_map and ("cargo" in local_map or "abono" in local_map):
                header_row_idx = idx
                col_map = local_map
                break

        if header_row_idx is None:
            continue

        for row in rows[header_row_idx + 1:]:
            if row is None or all(c is None for c in row):
                continue
            fecha_val = row[col_map["fecha"]] if "fecha" in col_map else None
            if fecha_val is None:
                continue
            desc_val = row[col_map["descripcion"]] if "descripcion" in col_map else ""
            cargo_val = row[col_map["cargo"]] if "cargo" in col_map else None
            abono_val = row[col_map["abono"]] if "abono" in col_map else None
            saldo_val = row[col_map["saldo"]] if "saldo" in col_map else None

            def to_float(v):
                if v is None or v == "":
                    return 0.0
                if isinstance(v, (int, float)):
                    return float(v)
                return _parse_amount(str(v))

            fecha_str = fecha_val.strftime("%d/%m/%Y") if hasattr(fecha_val, "strftime") else str(fecha_val)

            transactions.append(
                Transaction(
                    fecha=fecha_str,
                    descripcion=str(desc_val) if desc_val is not None else "",
                    cargo=to_float(cargo_val),
                    abono=to_float(abono_val),
                    saldo=str(saldo_val) if saldo_val is not None else "",
                )
            )

    if not transactions:
        result.advertencias.append(
            "No se detectaron movimientos en el Excel. Revisa que existan "
            "columnas reconocibles (Fecha, Descripción/Detalle, Cargo, Abono)."
        )

    result.transacciones = transactions
    return result


def parse_file(file_path_or_buffer, filename: str) -> ParseResult:
    lower = filename.lower()
    if lower.endswith(".pdf"):
        return parse_pdf(file_path_or_buffer)
    elif lower.endswith((".xlsx", ".xlsm", ".xls")):
        return parse_excel(file_path_or_buffer)
    else:
        raise ValueError("Formato de archivo no soportado (usa PDF o Excel).")
