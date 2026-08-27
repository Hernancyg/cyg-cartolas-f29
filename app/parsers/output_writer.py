"""
Genera el Excel de salida replicando el formato de la plantilla ExportBanco.xlsx:

  - Hoja llamada "Banco"
  - Columna A vacía (ancho 13)
  - Columnas B:E con encabezado en fila 1:
        B: FECHA DIA/MES
        C: DETALLE DE TRANSACCION
        D: MONTO CHEQUES O CARGOS
        E: MONTO DEPOSITOS O ABONOS
  - Fuente Arial 8 en todo, encabezado en negrita y centrado
  - Bordes finos en todas las celdas de la tabla (B:E)
  - Formato de fecha dd-mm-yyyy en columna B, centrado
  - Formato numérico #,##0 en columna D
  - Formato contable en columna E
"""

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADERS = [
    "FECHA DIA/MES",
    "DETALLE DE TRANSACCION",
    "MONTO CHEQUES O CARGOS",
    "MONTO DEPOSITOS O ABONOS",
]

COL_WIDTHS = {"A": 13.0, "B": 13.0, "C": 32.63, "D": 13.36, "E": 12.36}
NUMFMT_CARGO = "#,##0"
NUMFMT_ABONO = '_ * #,##0_ ;_ * \\-#,##0_ ;_ * \\-_ ;_ @_ '
NUMFMT_FECHA = "dd-mm-yyyy"

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _parse_fecha(fecha_str):
    """Intenta convertir 'dd/mm/aaaa' o 'dd-mm-aaaa' a datetime; si falla,
    devuelve el string original para no perder el dato."""
    if not fecha_str:
        return ""
    fecha_str = fecha_str.strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.strptime(fecha_str, fmt)
        except ValueError:
            continue
    return fecha_str


def build_output_workbook(transactions):
    """transactions: lista de bank_parsers.Transaction"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Banco"

    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    header_font = Font(name="Arial", size=8, bold=True)
    body_font = Font(name="Arial", size=8, bold=False)
    center = Alignment(horizontal="center")

    for i, text in enumerate(HEADERS):
        col_letter = get_column_letter(2 + i)  # B,C,D,E
        cell = ws[f"{col_letter}1"]
        cell.value = text
        cell.font = header_font
        cell.alignment = center
        cell.border = BORDER

    row_idx = 2
    for tx in transactions:
        fecha_val = _parse_fecha(tx.fecha)

        b = ws[f"B{row_idx}"]
        b.value = fecha_val
        b.font = body_font
        b.alignment = center
        b.border = BORDER
        if isinstance(fecha_val, datetime):
            b.number_format = NUMFMT_FECHA

        c = ws[f"C{row_idx}"]
        c.value = tx.descripcion
        c.font = body_font
        c.border = BORDER

        d = ws[f"D{row_idx}"]
        d.font = body_font
        d.number_format = NUMFMT_CARGO
        d.border = BORDER
        if tx.cargo:
            d.value = tx.cargo

        e = ws[f"E{row_idx}"]
        e.font = body_font
        e.number_format = NUMFMT_ABONO
        e.border = BORDER
        if tx.abono:
            e.value = tx.abono

        row_idx += 1

    return wb
